import time
import datetime
import os
import sys
import logging
import subprocess
import socket
import shutil
import tempfile
from dotenv import load_dotenv
import undetected_chromedriver as uc
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import StaleElementReferenceException, NoSuchWindowException, WebDriverException, ElementClickInterceptedException, NoSuchElementException
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
import platform
import requests
import re
import datetime
import random
import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Import configuration and ICD9 codes
from config import *
from icd9_codes import icd9_substitutes
from upload import upload, setup_driver as upload_setup_driver, cleanup_brave as upload_cleanup_brave

def debug_page_state(driver, reason="Debug"):
    """Save screenshot and page info for debugging"""
    try:
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Check browser connection first
        if not check_browser_connection(driver):
            print("Browser connection lost, skipping debug capture")
            return
        
        # Take screenshot safely
        screenshot_path = f"debug_screenshot_{timestamp}_{reason}.png"
        if safe_screenshot(driver, screenshot_path):
            print(f"Debug screenshot saved: {screenshot_path}")
        
        # Save page source snippet
        try:
            page_source = driver.page_source[:2000] + "..." if len(driver.page_source) > 2000 else driver.page_source
            source_path = f"debug_source_{timestamp}_{reason}.txt"
            with open(source_path, 'w', encoding='utf-8') as f:
                f.write(f"URL: {driver.current_url}\n")
                f.write(f"Title: {driver.title}\n")
                f.write(f"Window handles: {len(driver.window_handles)}\n")
                f.write("="*50 + "\n")
                f.write(page_source)
            print(f"Debug source saved: {source_path}")
        except Exception as e:
            print(f"Failed to save page source: {e}")
        
    except Exception as e:
        print(f"Debug save failed: {e}")

def check_browser_connection(driver):
    """Check if browser connection is still valid"""
    try:
        # Try to get current URL to test connection
        current_url = driver.current_url
        return True
    except Exception as e:
        print(f"Browser connection lost: {e}")
        return False

def safe_screenshot(driver, filename):
    """Take screenshot with error handling"""
    try:
        driver.save_screenshot(filename)
        return True
    except Exception as e:
        print(f"Screenshot failed: {e}")
        return False

def safe_close_extra_windows(driver, main_window):
    """Fastest possible window cleanup - minimal operations"""
    try:
        # Get current handles once
        handles = driver.window_handles
        
        # If only one window, we're done
        if len(handles) <= 1:
            return True
            
        # Close extra windows with minimal error checking
        for handle in handles:
            if handle != main_window:
                try:
                    driver.switch_to.window(handle)
                    driver.close()
                except:
                    pass
        
        # Switch to main window once at the end
        if main_window in driver.window_handles:
            driver.switch_to.window(main_window)
        elif driver.window_handles:
            driver.switch_to.window(driver.window_handles[0])
        
        return True
    except:
        return False

# Advanced Anti-Detection Configuration
class BrowserFingerprint:
    """Generate realistic browser fingerprints to avoid detection"""
    
    @staticmethod
    def get_random_user_agent():
        """Return a random realistic user agent"""
        user_agents = [
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36", 
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36"
        ]
        return random.choice(user_agents)
    
    @staticmethod
    def get_viewport_size():
        """Return a random realistic viewport size"""
        viewports = [
            (1920, 1080), (1366, 768), (1440, 900), (1536, 864), (1680, 1050)
        ]
        return random.choice(viewports)
    
    @staticmethod
    def get_timezone():
        """Return a realistic timezone"""
        timezones = ["America/New_York", "America/Chicago", "America/Denver", "America/Los_Angeles", "America/Vancouver"]
        return random.choice(timezones)

# Load the environment variables
load_dotenv()

# Set up logging for failed ICD9 codes only
# Create a custom logger to avoid capturing other library logs
icd9_logger = logging.getLogger('failed_icd9')
icd9_logger.setLevel(logging.INFO)

# Create file handler
file_handler = logging.FileHandler('failed_icd9_codes.log')
file_handler.setLevel(logging.INFO)

# Create formatter
formatter = logging.Formatter('%(asctime)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
file_handler.setFormatter(formatter)

# Add handler to logger
icd9_logger.addHandler(file_handler)

# Prevent propagation to root logger to avoid other logs
icd9_logger.propagate = False

# webdriver options
options = webdriver.ChromeOptions()
if headless_mode:
    options.add_argument("--headless")

# Add additional Chrome options for better stability and compatibility
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--disable-gpu")
options.add_argument("--disable-web-security")
options.add_argument("--disable-features=VizDisplayCompositor")
options.add_argument("--allow-running-insecure-content")
options.add_argument("--disable-features=VizDisplayCompositor,VizServiceDisplay")
options.add_argument("--disable-cors")
options.add_argument("--disable-site-isolation-trials")
options.add_argument("--disable-features=BlockInsecurePrivateNetworkRequests")
options.add_argument("--ignore-certificate-errors")
options.add_argument("--ignore-ssl-errors")
options.add_argument("--ignore-certificate-errors-spki-list")
options.add_argument("--ignore-ssl-errors-ignore-cert-validity")
options.add_argument("--allow-running-insecure-content")
options.add_argument("--disable-extensions")
options.add_argument("--disable-plugins")
options.add_argument("--disable-images")

# Specify the path to the Chrome binary if running on a Mac
if platform.system() == "Darwin":
    options.binary_location = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"

# Holds the Brave subprocess started by setup_chrome_driver so it can be
# terminated cleanly when the bot finishes (prevents orphaned processes).
_brave_process = None


def cleanup_brave():
    """Terminate the bot-owned Brave subprocess if it is still running."""
    global _brave_process
    if _brave_process is not None:
        try:
            _brave_process.terminate()
            _brave_process.wait(timeout=5)
        except Exception:
            try:
                _brave_process.kill()
            except Exception:
                pass
        _brave_process = None


def setup_chrome_driver():
    """Start Brave with remote debugging and connect via regular ChromeDriver"""
    global _brave_process

    print("BRAVE BROWSER SETUP")
    print("=" * 50)

    try:
        # Step 1: Kill only a previously bot-owned Brave instance (scoped to the
        # bot's own user-data-dir so normal Brave windows are not affected).
        print("Step 1: Closing any previous bot-owned browser instance...")
        brave_clean_dir_check = os.path.join(tempfile.gettempdir(), "brave_clean")
        if sys.platform == "win32":
            os.system(f'taskkill /F /FI "WINDOWTITLE eq brave_clean*" >nul 2>&1')
        else:
            os.system(f"pkill -f 'user-data-dir={brave_clean_dir_check}' 2>/dev/null")
        time.sleep(1)
        
        # Step 2: Start Brave with remote debugging (ONLY method that avoids 403)
        if sys.platform == "win32":
            brave_candidates = [
                os.path.join(os.environ.get("LOCALAPPDATA", ""), "BraveSoftware", "Brave-Browser", "Application", "brave.exe"),
                r"C:\Program Files\BraveSoftware\Brave-Browser\Application\brave.exe",
                r"C:\Program Files (x86)\BraveSoftware\Brave-Browser\Application\brave.exe",
            ]
            brave_path = next((p for p in brave_candidates if os.path.exists(p)), None)
        else:
            brave_path = "/Applications/Brave Browser.app/Contents/MacOS/Brave Browser"
            if not os.path.exists(brave_path):
                brave_path = None
        
        if not brave_path:
            raise Exception("Brave Browser not found. Please install Brave Browser for 403 protection.")
        
        print("Step 2: Starting Brave locally (clean instance)...")
        
        # Clean up any existing debug directory
        brave_clean_dir = os.path.join(tempfile.gettempdir(), "brave_clean")
        if os.path.exists(brave_clean_dir):
            shutil.rmtree(brave_clean_dir, ignore_errors=True)
        brave_clean_dir_check = brave_clean_dir  # keep name consistent
        
        # Start Brave as clean as possible - like a user opened it
        brave_args = [
            brave_path,
            "--remote-debugging-port=9222",
            f"--user-data-dir={brave_clean_dir}",
            "--window-size=1920,1080",
        ]

        # Headless mode: use --headless=new (Chrome 112+) which avoids the 403
        # detection issues caused by the old --headless flag
        if headless_mode:
            brave_args.append("--headless=new")
            brave_args.append("--disable-gpu")
            print("  (headless=new mode enabled)")
        
        try:
            _brave_process = subprocess.Popen(brave_args, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            print("✓ Clean Brave instance started")
        except Exception as e:
            raise Exception(f"Failed to start Brave: {str(e)}")
        
        # Wait for Brave to be ready (check remote debugging port)
        print("Waiting for Brave to initialize...")
        max_attempts = 10
        for attempt in range(max_attempts):
            try:
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.settimeout(1)
                result = sock.connect_ex(('127.0.0.1', 9222))
                sock.close()
                if result == 0:
                    print("✓ Brave remote debugging ready")
                    break
            except:
                pass
            
            if attempt == max_attempts - 1:
                raise Exception("Brave failed to start remote debugging within 10 seconds")
            
            time.sleep(1)
            print(f"  Attempt {attempt + 1}/{max_attempts}...")
        
        # Step 3: Connect via REGULAR ChromeDriver (not undetected)
        print("Step 3: Connecting to Brave via ChromeDriver...")
        
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
        
        # Only use compatible options for this ChromeDriver version
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_argument("--disable-extensions")
        
        # Get ChromeDriver path
        driver_path = ChromeDriverManager().install()
        
        if "THIRD_PARTY_NOTICES" in driver_path:
            driver_dir = os.path.dirname(driver_path)
            chromedriver_name = "chromedriver.exe" if sys.platform == "win32" else "chromedriver"
            actual_driver = os.path.join(driver_dir, chromedriver_name)
            driver_path = actual_driver
        
        if not os.path.exists(driver_path):
            raise Exception(f"ChromeDriver not found at {driver_path}")
        
        if sys.platform != "win32" and not os.access(driver_path, os.X_OK):
            print("✓ Fixing ChromeDriver permissions...")
            os.chmod(driver_path, 0o755)
        
        # Use regular ChromeDriver with Brave remote debugging
        service = ChromeService(driver_path)
        driver = webdriver.Chrome(service=service, options=chrome_options)
        
        print("✓ Connected to Brave via ChromeDriver")
        
        # Set timeouts to prevent hanging
        driver.set_page_load_timeout(30)
        driver.implicitly_wait(10)
        
        # Hide webdriver property for additional stealth
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        
        print("✓ Clean local Brave ready!")
        
        return driver
        
    except Exception as e:
        print(f"✗ Brave browser setup failed: {str(e)}")
        print("\nMANUAL SETUP INSTRUCTIONS:")
        print("1. Close all browsers")
        print("2. Install Brave Browser if not installed")
        print("3. Try running the bot again")
        print("   Or manually start Brave with:")
        if sys.platform == "win32":
            print(r'   "C:\Program Files\BraveSoftware\Brave-Browser\Application\brave.exe" --remote-debugging-port=9222 --user-data-dir=%TEMP%\brave_clean')
        else:
            print('   /Applications/Brave\\ Browser.app/Contents/MacOS/Brave\\ Browser --remote-debugging-port=9222 --user-data-dir=/tmp/brave_clean')
        
        raise WebDriverException(f"Brave setup failed: {str(e)}")

def ping_dasrecord(message):
    """Send status updates to DasRecord or other systems"""
    try:
        # Implementation for external status updates if needed
        pass
    except:
        pass

def export_to_excel(billing_data):
    """Export billing data to Excel file with specified columns"""
    excel_filename = f"{billing_year}{billing_month.zfill(2)}{billing_day.zfill(2)}_billing_export.xlsx"
    
    # Column headers in the required order
    headers = [
        "date_of_birth", "last_name", "first_name", "PHN", 
        "date_of_service", "doc_last_name", "doc_first_name", "billing_item", "diagnosis", 
        "location", "province"
    ]
    
    try:
        # Try to load existing workbook
        if os.path.exists(excel_filename):
            workbook = load_workbook(excel_filename)
            sheet = workbook.active
        else:
            # Create new workbook with headers
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Billing Data"
            # Add headers
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)
        
        # Find the next empty row
        next_row = sheet.max_row + 1
        
        # Add the billing data
        row_data = [
            billing_data.get("date_of_birth", ""),
            billing_data.get("last_name", ""), 
            billing_data.get("first_name", ""),
            billing_data.get("PHN", ""),
            billing_data.get("date_of_service", ""),
            billing_data.get("doc_last_name", ""),  # Hardcoded as requested
            billing_data.get("doc_first_name", ""),  # Hardcoded as requested
            billing_data.get("billing_item", ""),
            billing_data.get("diagnosis", ""),
            "V",  # Hardcoded as requested
            billing_data.get("province", "")
        ]
        
        # Write data to the row
        for col_num, value in enumerate(row_data, 1):
            sheet.cell(row=next_row, column=col_num, value=value)
        
        # Save the workbook
        workbook.save(excel_filename)
        print(f"✅ Exported billing data to {excel_filename}, row {next_row}")
        return True
        
    except Exception as e:
        print(f"❌ Error exporting to Excel: {e}")
        return False

def update_appointment_status(driver, day_sheet_window, appointment_element=None):
    """Update appointment status to 'B' (Billed) via the specific appointment's link"""
    try:
        print("🔄 Updating appointment status to 'B'...")
        
        # Ensure we're in the day sheet window
        if day_sheet_window in driver.window_handles:
            driver.switch_to.window(day_sheet_window)
        else:
            driver.switch_to.window(driver.window_handles[0])
            day_sheet_window = driver.window_handles[0]
        
        # Find the appointment link WITHIN the specific appointment element (not the entire page)
        appt_link = None
        
        # Try to use the provided appointment element first
        if appointment_element:
            try:
                appt_link = appointment_element.find_element(By.CSS_SELECTOR, "a.apptLink")
                link_text = appt_link.text.strip()
                print(f"🎯 Found appointment link for: '{link_text}'")
            except (StaleElementReferenceException, NoSuchElementException):
                print("⚠️ Appointment element is stale or doesn't have apptLink, refreshing...")
                appt_link = None
        
        # Fallback: if no element provided or element is stale, get first appointment link
        if not appt_link:
            try:
                appt_link = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "a.apptLink"))
                )
                print("🎯 Found appointment link (fallback search)")
            except:
                print("❌ Could not find any appointment link")
                return False
        
        # Scroll and click the specific appointment link
        try:
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", appt_link)
            time.sleep(0.5)
            appt_link.click()
            print("✅ Clicked appointment link")
        except Exception as click_error:
            print(f"⚠️ Normal click failed, trying JavaScript: {click_error}")
            try:
                driver.execute_script("arguments[0].click();", appt_link)
                print("✅ Clicked appointment link via JavaScript")
            except Exception as js_error:
                print(f"❌ Click failed: {js_error}")
                return False
        
        # Wait for new window to open  
        WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) > 1)
        
        # Switch to the new appointment window
        for handle in driver.window_handles:
            if handle != day_sheet_window:
                driver.switch_to.window(handle)
                break
        
        print("📝 Opened appointment details window")
        
        # Wait for page to load and find the status select element
        status_select_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.NAME, "status"))
        )
        
        # Create Select object and choose option 'B' (value="B" label="Billed")
        status_select = Select(status_select_element)
        status_select.select_by_value("B")
        print("✅ Changed status to 'B' (Billed)")
        
        # Find and click the update button (simplified since structure is consistent)
        print("🔍 Looking for Update Appt button...")
        try:
            update_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "input[value='Update Appt']"))
            )
            print("✅ Found Update Appt button")
        except:
            print("❌ Could not find Update Appt button")
            debug_page_state(driver, "update_button_not_found")
            return False
        
        # Click the update button
        print("🔄 Clicking Update Appt button...")
        try:
            # Scroll to button and click
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", update_button)
            time.sleep(0.5)
            update_button.click()
            print("✅ Update Appt button clicked successfully")
        except Exception as click_error:
            print(f"⚠️ Normal click failed, trying JavaScript: {click_error}")
            try:
                driver.execute_script("arguments[0].click();", update_button)
                print("✅ Update Appt button clicked via JavaScript")
            except Exception as js_error:
                print(f"❌ Button click failed: {js_error}")
                return False
        
        # onButUpdate() in the onclick handler closes the window via JS.
        # Wait up to 5 s for it to self-close; only close manually if it doesn't.
        try:
            WebDriverWait(driver, 5).until(lambda d: len(d.window_handles) == 1)
            print("✅ Appointment window closed by EMR")
        except Exception:
            # Window didn't self-close — close it ourselves
            try:
                driver.close()
            except Exception:
                pass

        driver.switch_to.window(day_sheet_window)
        print("✅ Updated appointment status and returned to day sheet")

        return True
        
    except Exception as e:
        print(f"❌ Error updating appointment status: {e}")
        # Try to get back to day sheet window
        try:
            driver.switch_to.window(day_sheet_window)
        except:
            pass
        return False

def ping_dasrecord(message):
    try:
        payload = {'text': message}
        response = requests.post("https://relayproxy.vercel.app/das_record_slack", json=payload)
        if response.status_code == 200:
            print("Successfully sent message to DAS Record.")
        else:
            print("Failed to send message to DAS Record.")
    except requests.exceptions.RequestException as e:
        print(f"An error occurred while sending message to DAS Record: {e}")

def login_to_oscar(driver):
    """Handle EMR login and navigate to provider schedule automatically"""
    
    print("\n" + "="*50)
    print("EMR LOGIN & NAVIGATION")
    print("="*50)
    
    try:
        # Navigate to the EMR login page
        print("🌐 Navigating to EMR login page...")
        emr_url = "https://well-kerrisdale.kai-oscar.com/oscar"
        
        # Set shorter timeout for this connection test
        original_timeout = 30
        driver.set_page_load_timeout(15)  # Shorter timeout to prevent hanging
        
        max_attempts = 3
        connected = False
        
        for attempt in range(max_attempts):
            try:
                print(f"  Connection attempt {attempt + 1}/{max_attempts}...")
                
                driver.get(emr_url)
                
                # Wait for page to load
                WebDriverWait(driver, 10).until(
                    lambda d: d.execute_script("return document.readyState") == "complete"
                )
                
                current_url = driver.current_url
                print(f"✓ Connected to: {current_url}")
                connected = True
                break
                
            except Exception as e:
                print(f"  Attempt {attempt + 1} failed: {str(e)[:100]}...")
                
                if attempt == max_attempts - 1:
                    print("❌ EMR website appears to be down or very slow")
                    print("Please try again later or check EMR site status")
                    return False
                else:
                    time.sleep(2)  # Wait before retry
        
        # Restore normal timeout
        driver.set_page_load_timeout(original_timeout)
        
        if not connected:
            return False
        
        # Check current page status
        current_url = driver.current_url
        page_source = driver.page_source.lower() if driver.page_source else ""
        
        print(f"Current URL: {current_url}")
        
        # Check if already on provider schedule page (already logged in)
        if "providercontrol" in current_url and "provider" in current_url:
            print("✓ Already logged in and on provider schedule!")
            return True
        
        # Check if need to login
        if any(indicator in page_source for indicator in ["username", "password", "login", "sign in"]):
            print("🔐 Login form detected - attempting automatic login...")
            
            # Get credentials from environment
            username = os.getenv('OSCAR_USERNAME')
            password = os.getenv('OSCAR_PASSWORD')
            
            if not username or not password:
                print("❌ No credentials found in environment variables")
                return False
                
            try:
                # Find username field
                username_field = None
                username_selectors = [
                    "input[name='username']", "input[name='user']", "input[name='login']",
                    "input[id='username']", "input[id='user']", "input[type='text']:first-of-type"
                ]
                
                for selector in username_selectors:
                    try:
                        username_field = driver.find_element(By.CSS_SELECTOR, selector)
                        print(f"✓ Found username field: {selector}")
                        break
                    except:
                        continue
                
                # Find password field
                password_field = None
                password_selectors = [
                    "input[name='password']", "input[name='pass']", 
                    "input[id='password']", "input[type='password']"
                ]
                
                for selector in password_selectors:
                    try:
                        password_field = driver.find_element(By.CSS_SELECTOR, selector)
                        print(f"✓ Found password field: {selector}")
                        break
                    except:
                        continue
                
                if not username_field or not password_field:
                    print("❌ Could not find login form fields")
                    return False
                
                # Enter credentials
                print("🔑 Entering credentials...")
                username_field.clear()
                username_field.send_keys(username)
                print("✓ Username entered")
                
                password_field.clear()
                password_field.send_keys(password)
                print("✓ Password entered")
                
                # Find and click login button
                login_button = None
                login_selectors = [
                    "input[type='submit']", "button[type='submit']", 
                    "input[value*='Login']", "input[value*='Sign']"
                ]
                
                for selector in login_selectors:
                    try:
                        login_button = driver.find_element(By.CSS_SELECTOR, selector)
                        print(f"✓ Found login button: {selector}")
                        break
                    except:
                        continue
                
                if login_button:
                    login_button.click()
                    print("✓ Login button clicked")
                else:
                    # Try submitting with Enter key
                    password_field.send_keys(Keys.RETURN)
                    print("✓ Login submitted with Enter key")
                
                # Wait for login to process
                print("⏳ Waiting for login to process...")
                WebDriverWait(driver, 10).until(lambda d: d.current_url != current_url)
                
                # Check if login was successful
                new_url = driver.current_url
                print(f"After login URL: {new_url}")
                
                if new_url != current_url and "login" not in new_url.lower():
                    print("✓ Login successful!")
                else:
                    print("❌ Login may have failed")
                    return False
                    
            except Exception as login_error:
                print(f"❌ Login automation failed: {login_error}")
                return False
        
        # Navigate to provider schedule for today's date
        print("🗓️ Navigating to today's provider schedule...")
        billing_date = f"https://well-kerrisdale.kai-oscar.com/oscar/provider/providercontrol.jsp?year={billing_year}&month={billing_month}&day={billing_day}&view=0&displaymode=day&dboperation=searchappointmentday&viewall=0"
        
        # Use shorter timeout for navigation
        driver.set_page_load_timeout(15)
        
        max_nav_attempts = 2
        for nav_attempt in range(max_nav_attempts):
            try:
                print(f"  Navigation attempt {nav_attempt + 1}/{max_nav_attempts}...")
                
                driver.get(billing_date)
                time.sleep(0.5)
                
                final_url = driver.current_url
                print(f"Final URL: {final_url}")
                
                if "providercontrol" in final_url:
                    print("✓ Successfully navigated to provider schedule!")
                    # Restore normal timeout
                    driver.set_page_load_timeout(original_timeout)
                    return True
                else:
                    print("⚠️ May not be on the correct provider schedule page")
                    # Restore normal timeout
                    driver.set_page_load_timeout(original_timeout)
                    return True  # Continue anyway
                    
            except Exception as nav_error:
                print(f"  Navigation attempt {nav_attempt + 1} failed: {str(nav_error)[:100]}...")
                if nav_attempt == max_nav_attempts - 1:
                    print("❌ Navigation timeout - EMR may be slow")
                    # Restore normal timeout
                    driver.set_page_load_timeout(original_timeout)
                    return False
                else:
                    time.sleep(2)  # Wait before retry
            
    except Exception as e:
        print(f"❌ Login/navigation error: {str(e)}")
        return False

def navigate_to_billing_date(driver):
    billing_date = f"https://well-kerrisdale.kai-oscar.com/oscar/provider/providercontrol.jsp?year={billing_year}&month={billing_month}&day={billing_day}&view=0&displaymode=day&dboperation=searchappointmentday&viewall=0"
    print(f"Navigating to: {billing_date}")
    
    # Navigate to the billing date URL
    driver.get(billing_date)
    
    # Wait for the page to load and check if we're on the right page
    WebDriverWait(driver, long_delay * 2).until(
        lambda driver: "providercontrol.jsp" in driver.current_url
    )
    
    print(f"Successfully navigated to: {driver.current_url}")
    
    # Wait for the appointment schedule to load
    WebDriverWait(driver, long_delay).until(
        EC.presence_of_element_located((By.TAG_NAME, "body"))
    )
    
    # Additional wait for dynamic content to load
    time.sleep(0.3)

def get_appointments(driver):
    # Check browser connection first
    if not check_browser_connection(driver):
        print("❌ Browser connection lost in get_appointments")
        return []
    
    # Wait for appointments to load, then return them
    try:
        WebDriverWait(driver, long_delay).until(
            EC.presence_of_element_located((By.CLASS_NAME, "appt"))
        )
        appointments = driver.find_elements(By.CLASS_NAME, "appt")
        print(f"✅ Found {len(appointments)} appointments")
        return appointments
    except Exception as e:
        # If no appointments found, return empty list
        print(f"⚠️  No appointments found or appointments not loaded yet: {e}")
        return []

def extract_patient_info(driver):
    """Extract patient information from the billing form page using exact table structure"""
    patient_info = {}
    
    try:
        print("🔍 Extracting patient information from billing form table structure...")
        
        # Find the Patient Information table using multiple fallback strategies
        patient_table = None
        patient_table_strategies = [
            # td with direct text
            "//td[contains(text(), 'Patient Information')]/ancestor::table[1]",
            # th with direct text
            "//th[contains(text(), 'Patient Information')]/ancestor::table[1]",
            # any element with nested text (e.g. <td><b>Patient Information</b></td>)
            "//*[contains(., 'Patient Information')][self::td or self::th]/ancestor::table[1]",
            # normalize-space to handle extra whitespace
            "//td[contains(normalize-space(.), 'Patient Information')]/ancestor::table[1]",
            "//th[contains(normalize-space(.), 'Patient Information')]/ancestor::table[1]",
        ]
        for strategy in patient_table_strategies:
            try:
                patient_table = driver.find_element(By.XPATH, strategy)
                print(f"✅ Found patient table using: {strategy}")
                break
            except NoSuchElementException:
                continue
        if patient_table is None:
            raise NoSuchElementException("Could not locate Patient Information table with any known selector")
        
        # Extract Patient Name (2nd cell, 1st data row) - Format: "LASTNAME,FIRSTNAME"
        try:
            patient_name_xpath = ".//tr[2]/td[2]"  # First data row, second cell
            patient_name_element = patient_table.find_element(By.XPATH, patient_name_xpath)
            patient_name = patient_name_element.text.strip()
            
            if ',' in patient_name:
                parts = patient_name.split(',')
                patient_info["last_name"] = parts[0].strip()
                patient_info["first_name"] = parts[1].strip() if len(parts) > 1 else ""
            else:
                # If no comma, assume it's all one name
                patient_info["last_name"] = patient_name
                patient_info["first_name"] = ""
            
            print(f"✅ Found Patient Name: {patient_info['last_name']}, {patient_info['first_name']}")
        except Exception as e:
            print(f"⚠️ Could not extract patient name: {e}")
            patient_info["last_name"] = ""
            patient_info["first_name"] = ""
        
        # Extract Patient PHN (4th cell, 1st data row)
        try:
            phn_xpath = ".//tr[2]/td[4]"  # First data row, fourth cell
            phn_element = patient_table.find_element(By.XPATH, phn_xpath)
            patient_info["phn"] = phn_element.text.strip()
            print(f"✅ Found PHN: {patient_info['phn']}")
        except Exception as e:
            print(f"⚠️ Could not extract PHN: {e}")
            patient_info["phn"] = ""
        
        # Extract Date of Birth (2nd cell, 2nd data row) - Format: YYYYMMDD
        try:
            dob_xpath = ".//tr[3]/td[2]"  # Second data row, second cell
            dob_element = patient_table.find_element(By.XPATH, dob_xpath)
            dob_raw = dob_element.text.strip()
            
            # Convert YYYYMMDD to YYYY-MM-DD
            if len(dob_raw) == 8 and dob_raw.isdigit():
                formatted_dob = f"{dob_raw[:4]}-{dob_raw[4:6]}-{dob_raw[6:8]}"
                patient_info["date_of_birth"] = formatted_dob
                print(f"✅ Found DOB: {patient_info['date_of_birth']}")
            else:
                patient_info["date_of_birth"] = dob_raw
                print(f"✅ Found DOB (raw): {patient_info['date_of_birth']}")
        except Exception as e:
            print(f"⚠️ Could not extract DOB: {e}")
            patient_info["date_of_birth"] = ""
        
        # Extract Province/Health Card Type (6th cell, 1st data row)
        try:
            province_xpath = ".//tr[2]/td[6]"  # First data row, sixth cell
            province_element = patient_table.find_element(By.XPATH, province_xpath)
            patient_info["province"] = province_element.text.strip()
            print(f"✅ Found Province: {patient_info['province']}")
        except Exception as e:
            print(f"⚠️ Could not extract province: {e}")
            patient_info["province"] = "BC"  # Default
        
        # Set hardcoded location
        patient_info["location"] = "V"
        print(f"✅ Set Location: {patient_info['location']}")
        
        # Extract billing/service code from the service table
        try:
            # Find the service code table using multiple fallback strategies
            service_table = None
            service_table_strategies = [
                "//td[contains(text(), 'Service Code')]/ancestor::table[1]",
                "//th[contains(text(), 'Service Code')]/ancestor::table[1]",
                "//*[contains(., 'Service Code')][self::td or self::th]/ancestor::table[1]",
                "//td[contains(normalize-space(.), 'Service Code')]/ancestor::table[1]",
                "//th[contains(normalize-space(.), 'Service Code')]/ancestor::table[1]",
            ]
            for strategy in service_table_strategies:
                try:
                    service_table = driver.find_element(By.XPATH, strategy)
                    print(f"✅ Found service table using: {strategy}")
                    break
                except NoSuchElementException:
                    continue
            if service_table is None:
                raise NoSuchElementException("Could not locate Service Code table with any known selector")
            
            # Get the service code from first data row, first cell
            service_code_xpath = ".//tr[2]/td[1]"  # First data row, first cell
            service_code_element = service_table.find_element(By.XPATH, service_code_xpath)
            patient_info["billing_code"] = service_code_element.text.strip()
            print(f"✅ Found Billing Code: {patient_info['billing_code']}")
        except Exception as e:
            print(f"⚠️ Could not extract billing code: {e}")
            patient_info["billing_code"] = "A001A"  # Fallback default
        
        print(f"📊 Successfully extracted patient info: {patient_info}")
        return patient_info
        
    except Exception as e:
        print(f"❌ Error extracting patient info from table structure: {e}")
        return {
            "date_of_birth": "",
            "last_name": "",
            "first_name": "",
            "phn": "",
            "location": "V",
            "province": "BC",
            "billing_code": "A001A"  # Fallback default
        }

def format_date(date_str):
    """Convert date to YYYY-MM-DD format"""
    try:
        # Try different date formats
        formats = [
            "%Y-%m-%d",  # Already correct format
            "%m/%d/%Y",  # MM/DD/YYYY
            "%d/%m/%Y",  # DD/MM/YYYY
            "%m-%d-%Y",  # MM-DD-YYYY
            "%d-%m-%Y"   # DD-MM-YYYY
        ]
        
        for fmt in formats:
            try:
                date_obj = datetime.datetime.strptime(date_str, fmt)
                return date_obj.strftime("%Y-%m-%d")
            except:
                continue
                
        return date_str  # Return as-is if no format matches
    except:
        return date_str

def process_appointment(driver, appointment, day_sheet_window):
    global cumulative_end_time
    global counseling_appointment_count
    encounter_window = None  # Initialize at function level
    billing_window = None   # Initialize at function level
    
    # Check browser connection before proceeding
    if not check_browser_connection(driver):
        print("❌ Browser connection lost before processing appointment")
        raise WebDriverException("Browser connection lost")
    
    try:
        # print(f"Processing appointment: {appointment.text}")
        appointment_status = appointment.find_element(By.XPATH, ".//img[1]").get_attribute("title")
    except Exception as e:
        print(f"❌ Failed to read appointment status: {e}")
        raise
    # print(f"Appointment status: {appointment_status}")

    if appointment_status in ["No Show","Billed/Verified","Billed/Signed", "Billed", "Cancelled"]:
        print(f"⏭️  Skipping - Status: {appointment_status}")
        return False

    if "Track,Fast" in appointment.text:
        print("🏃 Fast track appointment. Skipping.")
        return False

    print(f"📊 Processing appointment - Status: {appointment_status}")

    counseling = False
    e_chart = appointment.find_element(By.XPATH, ".//a[contains(@title, 'Encounter')]")
    billing_button = appointment.find_element(By.XPATH, ".//a[contains(@title, 'Billing')]")
    master_record_button = appointment.find_element(By.XPATH, ".//a[contains(@title, 'Master Record')]")
    prescriptions_button = appointment.find_element(By.XPATH, ".//a[contains(@title, 'Prescriptions')]")

    start_time = appointment.find_element(By.XPATH, "./preceding-sibling::td[@align='RIGHT'][1]/a[1]").get_attribute("title").split(" - ")[0]
    start_time = datetime.datetime.strptime(start_time, "%I:%M%p").strftime("%H:%M")
    current_start_time = datetime.datetime.strptime(start_time, "%H:%M")
    print(f"Start time: {start_time}")

    if cumulative_end_time and cumulative_end_time > current_start_time:
        current_start_time = cumulative_end_time
        print(f"Adjusted start time: {current_start_time.strftime('%H:%M')}")

    # Check browser connection before clicking encounter
    if not check_browser_connection(driver):
        print("❌ Browser connection lost before clicking encounter")
        raise WebDriverException("Browser connection lost")
    
    # Scroll to ensure the element is visible and click using JavaScript if needed
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", e_chart)
    time.sleep(0.2)
    
    print(f"🎡 Opening encounter window...")
    
    try:
        # Try normal click first
        WebDriverWait(driver, short_delay).until(EC.element_to_be_clickable(e_chart))
        e_chart.click()
        print("✅ Encounter button clicked")
    except (WebDriverException, ElementClickInterceptedException):
        # Check connection before trying JavaScript click
        if not check_browser_connection(driver):
            print("❌ Browser connection lost during click attempt")
            raise WebDriverException("Browser connection lost")
        # If normal click fails, use JavaScript click
        print("🔄 Normal click failed, trying JavaScript click...")
        driver.execute_script("arguments[0].click();", e_chart)
        print("✅ JavaScript click successful")

    # Wait for new window to open
    WebDriverWait(driver, long_delay).until(lambda d: len(d.window_handles) > 1)
    
    # Find the correct encounter window by checking URLs/content
    original_window = driver.current_window_handle
    encounter_window = None
    
    print(f"Total windows after click: {len(driver.window_handles)}")
    
    for handle in driver.window_handles:
        if handle != original_window:
            driver.switch_to.window(handle)
            current_url = driver.current_url
            print(f"Checking window with URL: {current_url}")
            
            # Skip old notes windows from previous patients
            if "method=allNotes" in current_url:
                print("Skipping old notes window")
                continue
            
            # Check if this is the encounter window by looking for encounter-specific keywords
            if any(keyword in current_url.lower() for keyword in ["encounter", "echart", "oscarencounter", "casemanagemententry"]):
                # Additional check: make sure it's NOT an old notes window
                if "method=setupmainencounter" in current_url.lower() or "setupmainencounter" in current_url.lower():
                    encounter_window = handle
                    print("Found encounter window by URL")
                    break
            
            # If URL doesn't help, check page content
            try:
                page_title = driver.title.lower()
                page_source_snippet = driver.page_source[:1000].lower()
                
                # Make sure it's not an old notes window
                if "encounter notes" in page_title:
                    print("Skipping old encounter notes window")
                    continue
                
                if any(keyword in page_title for keyword in ["encounter", "chart", "assessment"]) and "notes" not in page_title:
                    encounter_window = handle
                    print("Found encounter window by title")
                    break
                elif any(keyword in page_source_snippet for keyword in ["show all notes", "assessment", "plan", "progress note"]):
                    encounter_window = handle
                    print("Found encounter window by content")
                    break
                    
            except:
                continue
    
    if not encounter_window:
        # Fallback: use the last opened window
        print("Could not identify encounter window specifically, using last window")
        encounter_window = driver.window_handles[-1]
    
    driver.switch_to.window(encounter_window)
    print(f"Switched to encounter window: {driver.current_url}")

    # Quick encounter window processing
    _t0 = time.time()
    try:
        # Debug: Print page information
        print(f"Encounter window URL: {driver.current_url}")
        print(f"Encounter window title: {driver.title}")
        
        # Verify we're in the correct encounter window
        current_url = driver.current_url
        if "day" in current_url.lower() or "schedule" in current_url.lower():
            print("WARNING: Still appears to be in day sheet/schedule window!")
            debug_page_state(driver, "wrong_window_detected")
            return
        
        # Disable implicit wait so find_element/find_elements return immediately on miss
        # (implicit_wait=10 would burn 10s per failed selector otherwise)
        driver.implicitly_wait(0)
        _t1 = time.time(); print(f"[TIMING] encounter window ready: {_t1 - _t0:.2f}s")

        # Check for iframes that might contain the notes
        iframes = driver.find_elements(By.TAG_NAME, "iframe")
        _t2 = time.time(); print(f"[TIMING] iframe scan ({len(iframes)} found): {_t2 - _t1:.2f}s")
        print(f"Found {len(iframes)} iframes in encounter window")
        
        # Quick 403 check
        if "403" in driver.page_source[:300] or "Forbidden" in driver.page_source[:300]:
            driver.implicitly_wait(10)
            print("403 error detected")
            return
        
        # Find Show All Notes button — try instant find first, short wait only as fallback
        show_all_notes = None
        selectors = [
            "//input[@value='Show All Notes']",
            "//*[text()='Show All Notes']",
            "//*[contains(text(), 'Show All Notes')]",
            "//button[contains(text(), 'Show All')]",
            "//a[contains(text(), 'Show All')]",
            "//*[contains(text(), 'Show All')]",
        ]

        # Pass 1: instant find_element (page is already loaded)
        _t3 = time.time()
        for selector in selectors:
            try:
                show_all_notes = driver.find_element(By.XPATH, selector)
                print(f"Found Show All Notes button using: {selector}")
                break
            except NoSuchElementException:
                continue
        print(f"[TIMING] Pass 1 button search: {time.time() - _t3:.2f}s")

        # Pass 2: single short wait in case the DOM isn't quite settled
        if not show_all_notes:
            print("Button not immediately visible, waiting up to 3s...")
            _t4 = time.time()
            for selector in selectors:
                try:
                    show_all_notes = WebDriverWait(driver, 3).until(
                        EC.element_to_be_clickable((By.XPATH, selector))
                    )
                    print(f"Found Show All Notes button (delayed) using: {selector}")
                    break
                except:
                    continue
            print(f"[TIMING] Pass 2 button search: {time.time() - _t4:.2f}s")
        
        if not show_all_notes:
            print("Show All Notes button not found after all attempts")
            
            # Try searching in iframes
            iframes = driver.find_elements(By.TAG_NAME, "iframe")
            if iframes:
                print(f"Searching in {len(iframes)} iframes for Show All Notes button...")
                for i, iframe in enumerate(iframes):
                    try:
                        print(f"Switching to iframe {i+1}")
                        driver.switch_to.frame(iframe)
                        
                        # Try to find the button in this iframe
                        show_all_notes = driver.find_element(By.XPATH, "//*[contains(text(), 'Show All')]")
                        if show_all_notes:
                            print(f"Found Show All Notes button in iframe {i+1}")
                            break
                    except:
                        print(f"Button not found in iframe {i+1}")
                    finally:
                        # Switch back to main content
                        driver.switch_to.default_content()
                        driver.switch_to.window(driver.window_handles[1])  # Back to encounter window
            
            if not show_all_notes:
                debug_page_state(driver, "show_notes_not_found")
                # Try to find any button with 'Notes' in the text as a last resort
                try:
                    print("Searching for any button containing 'Notes'...")
                    show_all_notes = WebDriverWait(driver, 3).until(
                        EC.element_to_be_clickable((By.XPATH, "//*[contains(text(), 'Notes')]"))
                    )
                    print("Found alternative Notes button")
                except:
                    print("No Notes-related button found")
                    debug_page_state(driver, "no_notes_button_found")
                    return
        
        # Restore implicit wait before clicking (needed for subsequent find_element calls)
        driver.implicitly_wait(10)

        # Click the button
        _t_click = time.time()
        try:
            # Scroll to button and click
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", show_all_notes)
            time.sleep(0.1)
            show_all_notes.click()
            print(f"[TIMING] click Show All Notes: {time.time() - _t_click:.2f}s")
            print("Successfully clicked Show All Notes button")
        except Exception as click_error:
            print(f"Failed to click Show All Notes button: {click_error}")
            # Try JavaScript click as backup
            try:
                driver.execute_script("arguments[0].click();", show_all_notes)
                print("Successfully clicked using JavaScript")
            except:
                print("Both normal and JavaScript clicks failed")
                return
        
    except Exception as encounter_error:
        print(f"Encounter error: {str(encounter_error)}")
        return

    try:
        # Wait for new notes window to open (use a longer timeout — server can be slow)
        _t_win = time.time()
        WebDriverWait(driver, max(long_delay, 10)).until(lambda d: len(d.window_handles) > 2)
        print(f"[TIMING] notes window opened: {time.time() - _t_win:.2f}s")
        all_notes = driver.window_handles[2]
        driver.switch_to.window(all_notes)
        print("📄 Switched to all_notes window")

        # Wait for note content to load
        print("🔍 Reading patient notes...")
        _t_content = time.time()
        note_to_bill = WebDriverWait(driver, long_delay).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div[last()]"))
        )
        print(f"[TIMING] note content loaded: {time.time() - _t_content:.2f}s")
        note_content = note_to_bill.text
    except Exception as notes_error:
        print(f"❌ Error accessing notes window: {notes_error}")
        # Try to recover by going back to day sheet
        try:
            driver.switch_to.window(day_sheet_window)
        except:
            pass
        return

    print("📝 Extracting diagnosis from notes...")
    # print(f"📋 Full note content (first 500 chars): {note_content[:500]}...")
    
    if "A:" in note_content and "P:" in note_content:
        diagnosis = note_content.split("A:")[1].split("P:")[0]
        print(f"🔍 Extracted diagnosis section: {diagnosis.strip()}")
    else:
        print("❌ No A: and P: sections found, using full note content")
        diagnosis = note_content
    
    # Also search for ICD codes in the full note content as backup
    full_note_diagnosis = note_content

    def extract_diagnostic_code(diagnosis):
        # More comprehensive regex patterns for ICD-9 codes
        patterns = [
            r'ICD-?9:\s*([V]?\d{1,3}\.?\d{0,2})',  # Standard format: ICD9: 034.0 or ICD-9: V123.45
            r'ICD-?9\s*([V]?\d{1,3}\.?\d{0,2})',   # Without colon: ICD9 034.0
            r'\(ICD-?9:\s*([V]?\d{1,3}\.?\d{0,2})\)', # In parentheses: (ICD9: 034.0)
            r'ICD-?9\s*code:\s*([V]?\d{1,3}\.?\d{0,2})', # With "code": ICD9 code: 034.0
        ]
        
        for pattern in patterns:
            match = re.search(pattern, diagnosis, re.IGNORECASE)
            if match:
                print(f"🎯 ICD-9 pattern matched: {pattern} -> {match.group(1)}")
                return match.group(1)
        
        print(f"❌ No ICD-9 pattern matched in: {diagnosis[:100]}...")
        return None

    icd9_code = extract_diagnostic_code(diagnosis)
    
    # If no ICD code found in diagnosis section, try the full note content
    if not icd9_code:
        print("🔍 No ICD code in diagnosis section, searching full note content...")
        icd9_code = extract_diagnostic_code(full_note_diagnosis)
    
    icd9_code = icd9_code.replace(".", "") if icd9_code else "No ICD9 code found"
    print(f"🔍 Final extracted ICD-9 code: {icd9_code}")
    
    # Apply ICD9 code substitution if needed
    original_icd9 = icd9_code
    if icd9_code in icd9_substitutes:
        icd9_code = icd9_substitutes[icd9_code]
        print(f"🔄 Substituted {original_icd9} → {icd9_code}")

    appointment_was_processed = True  # track successful processing

    if "#C" in note_content:
        appointment_length = counseling_appointment_length
        counseling_appointment_count += 1
        counseling = True
    else:
        appointment_length = standard_appointment_length

    end_time = (current_start_time + datetime.timedelta(minutes=appointment_length)).strftime("%H:%M")
    cumulative_end_time = current_start_time + datetime.timedelta(minutes=appointment_length)
    print(f"End time: {end_time}")
    print(f"Cumulative end time: {cumulative_end_time.strftime('%H:%M')}")

    # Safely close windows and return to day sheet
    try:
        extra_window_count = len(driver.window_handles) - 1
        if extra_window_count > 0:
            print(f"🧹 Cleaning up {extra_window_count} extra windows...")
            
        success = safe_close_extra_windows(driver, day_sheet_window)
        if success:
            print("✅ Windows cleaned")
        else:
            print("⚠️ Window cleanup had issues")
                
    except Exception as window_error:
        print(f"Window cleanup error: {window_error}")
        # Try to get back to a valid window
        if driver.window_handles:
            driver.switch_to.window(driver.window_handles[0])

    # Check browser connection before billing operations
    if not check_browser_connection(driver):
        print("❌ Browser connection lost before billing window operations")
        return
    
    # Scroll to ensure billing button is visible and click using JavaScript if needed
    print(f"💵 Opening billing window...")
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", billing_button)
    time.sleep(0.2)
    
    try:
        # Try normal click first
        WebDriverWait(driver, short_delay).until(EC.element_to_be_clickable(billing_button))
        billing_button.click()
        print("✅ Billing button clicked")
    except (WebDriverException, ElementClickInterceptedException):
        # If normal click fails, use JavaScript click
        print("🔄 Normal billing click failed, trying JavaScript...")
        driver.execute_script("arguments[0].click();", billing_button)
    time.sleep(0.2)
    # Wait for billing window to open
    WebDriverWait(driver, long_delay).until(lambda d: len(d.window_handles) > 1)
    
    # Find the correct billing window by checking URLs/content
    current_window = driver.current_window_handle
    billing_window = None
    
    print(f"Total windows after billing click: {len(driver.window_handles)}")
    
    for handle in driver.window_handles:
        if handle != current_window:
            driver.switch_to.window(handle)
            current_url = driver.current_url
            print(f"Checking billing window with URL: {current_url}")
            
            # Check if this is the billing window by looking for billing-specific keywords
            if any(keyword in current_url.lower() for keyword in ["billing", "createbilling", "billingserviceform"]):
                billing_window = handle
                print("Found billing window by URL")
                break
            
            # If URL doesn't help, check page content
            try:
                page_title = driver.title.lower()
                page_source_snippet = driver.page_source[:1000].lower()
                
                if any(keyword in page_title for keyword in ["billing", "service"]):
                    billing_window = handle
                    print("Found billing window by title")
                    break
                elif any(keyword in page_source_snippet for keyword in ["billing form", "service code", "selectbillingform"]):
                    billing_window = handle
                    print("Found billing window by content")
                    break
                    
            except:
                continue
    
    if not billing_window:
        # Fallback: use the last opened window
        print("Could not identify billing window specifically, using last window")
        billing_window = driver.window_handles[-1]
    
    driver.switch_to.window(billing_window)
    print(f"Switched to billing window: {driver.current_url}")

    # Quick billing form processing
    try:
        # Wait for page to fully load
        print("⏳ Waiting for billing page to load...")
        WebDriverWait(driver, 10).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        
        print("🔍 Looking for billing form elements...")
        select_billing_form = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "selectBillingForm"))
        )
        select_billing_form.click()
        print("✅ Clicked selectBillingForm")

        # Select service code based on counseling
        service_code_type = "counseling" if counseling else "regular"
        service_code_xpath = "//*[@id='billingFormTable']/tbody/tr[1]/td[1]/table[1]/tbody/tr[3]/td[1]/label/input" if counseling else "//*[@id='billingFormTable']/tbody/tr[1]/td[1]/table[1]/tbody/tr[2]/td[1]/label/input"
        print(f"🎯 Selecting {service_code_type} service code...")
        service_code = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, service_code_xpath))
        )
        service_code.click()
        print(f"✅ Selected {service_code_type} service code")

        # Set times and diagnosis quickly
        print(f"⏰ Setting start time to {current_start_time.strftime('%H:%M')}...")
        start_time_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "serviceStartTime"))
        )
        start_time_input.send_keys(current_start_time.strftime("%H:%M"))
        print(f"✅ Start time set to: {current_start_time.strftime('%H:%M')}")
        
    except Exception as billing_form_error:
        print(f"Error in billing form processing: {billing_form_error}")
        debug_page_state(driver, "billing_form_error")
        return

    end_time_input = WebDriverWait(driver, 3).until(
        EC.presence_of_element_located((By.ID, "serviceEndTime"))
    )
    end_time_input.send_keys(end_time)

    diagnosis_input = WebDriverWait(driver, 3).until(
        EC.presence_of_element_located((By.ID, "billing_1_fee_dx1"))
    )
    diagnosis_input.send_keys(icd9_code)

    continue_button = WebDriverWait(driver, 3).until(
        EC.element_to_be_clickable((By.XPATH, "//*[@value='Continue']"))
    )
    continue_button.click()

    # Export mode logic - happens in the final billing confirmation window
    if export_mode:
        print("📊 Export mode enabled - extracting patient information...")
        try:
            # Wait for the final billing page to load (where patient info table is)
            WebDriverWait(driver, 10).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            
            # Extract patient information from the final billing confirmation page
            patient_data = extract_patient_info(driver)
            
            # Create billing data dictionary
            billing_data = {
                "date_of_birth": patient_data.get("date_of_birth", ""),
                "last_name": patient_data.get("last_name", ""),  # Match Excel header
                "first_name": patient_data.get("first_name", ""),  # Match Excel header
                "PHN": patient_data.get("phn", ""),
                "date_of_service": f"{billing_year}-{billing_month.zfill(2)}-{billing_day.zfill(2)}",
                "doc_last_name": "Das",  # Hardcoded
                "doc_first_name": "Prasenjit",  # Hardcoded
                "billing_item": patient_data.get("billing_code", ""),  # Match Excel header
                "diagnosis": icd9_code,  # Match Excel header
                "location": "V",  # Hardcoded location
                "province": patient_data.get("province", "BC")  # Default to BC
            }
            
            # Export to Excel
            export_success = export_to_excel(billing_data)
            
            if export_success:
                print("✅ Successfully exported to Excel")
                
                # Click Cancel instead of Save Bill
                try:
                    cancel_button = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, "//*[@value='Cancel' or @value='cancel' or contains(text(), 'Cancel')]"))
                    )
                    cancel_button.click()
                    print("✅ Clicked Cancel button (export mode)")
                except:
                    # If Cancel button not found, try to close window
                    print("⚠️ Cancel button not found, closing window")
                    driver.close()
                
                # Wait for billing window to close, then switch back to day sheet
                time.sleep(2)  # Give time for window to close
                
                # Switch back to day sheet and update the appointment status
                driver.switch_to.window(day_sheet_window)
                print("✅ Switched back to day sheet window")

                print("🔄 Starting appointment status update...")
                update_success = update_appointment_status(driver, day_sheet_window, appointment)

                if update_success:
                    print("✅ Appointment status updated successfully")
                else:
                    print("❌ Failed to update appointment status")
                
                return True  # Exit function for export mode
                
            else:
                print("❌ Export failed, continuing with normal processing")
                
        except Exception as export_error:
            print(f"❌ Export mode error: {export_error}")
            print("Falling back to normal processing")

    try:
        save_bill = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@value='Save Bill']"))
        )
        
        if not safe_mode:
            save_bill.click()
        else:
            print("Safe mode: Not saving")
            time.sleep(2)  # Reduced delay for safe mode
    except WebDriverException:
        print("Invalid billing code - logging for review")
        icd9_logger.info(f"Failed ICD9 code: {icd9_code} - Diagnosis: {diagnosis.strip()}")
        pass

    driver.switch_to.window(day_sheet_window)
    print("Switched back to day sheet window")
    return True

def process_appointments(driver, day_sheet_window):
    global cumulative_end_time
    global counseling_appointment_count

    cumulative_end_time = None
    counseling_appointment_count = 0

    # Get initial appointments and set runs based on count
    appointments = get_appointments(driver)
    total_appointments = len(appointments)
    if runs is not None:
        total_appointments = min(runs, total_appointments)
        print(f"Found {len(appointments)} appointments. Processing {total_appointments} as set by runs...")
    else:
        print(f"Found {total_appointments} appointments. Processing each one...")
    
    processed_count = 0
    appointment_index = 0  # walks all appointments; processed_count tracks only non-skipped ones
    
    # Process each appointment, handling stale element exceptions
    while processed_count < total_appointments:
        # Check browser connection before each appointment
        if not check_browser_connection(driver):
            print("❌ Browser connection lost during appointment processing")
            print(f"Successfully processed {processed_count} appointments before connection loss")
            break
        
        # Refresh appointment list in case of stale elements
        try:
            current_appointments = get_appointments(driver)
        except Exception as e:
            print(f"❌ Failed to refresh appointment list: {e}")
            break
        
        if appointment_index >= len(current_appointments):
            # No more appointments on the sheet
            break
            
        appointment = current_appointments[appointment_index]
        
        try:
            print(f"\nChecking appointment {appointment_index + 1} (processed {processed_count} of {total_appointments})")
            
            # Switch back to day sheet window before each appointment
            driver.switch_to.window(day_sheet_window)
            
            was_processed = process_appointment(driver, appointment, day_sheet_window)
            appointment_index += 1
            if was_processed:
                processed_count += 1
            
        except StaleElementReferenceException:
            print("⚠️  StaleElementReferenceException caught. Refreshing appointments and retrying...")
            # Don't advance index, retry same appointment
            continue
        except WebDriverException as we:
            print(f"❌ WebDriver error processing appointment {appointment_index + 1}: {str(we)}")
            
            # Check if it's a connection issue
            if not check_browser_connection(driver):
                print("🔌 Browser connection confirmed lost. Stopping processing.")
                break
            else:
                print("🔄 Browser still connected, skipping this appointment and continuing")
                appointment_index += 1
                processed_count += 1
                continue
        except Exception as e:
            print(f"❌ Error processing appointment {appointment_index + 1}: {str(e)}")
            appointment_index += 1
            processed_count += 1  # Count errored appointments so runs cap still works
            continue
    
    print(f"\nCompleted processing {processed_count} appointments!")
    if counseling_appointment_count > 0:
        print(f"Counseling appointments processed: {counseling_appointment_count}")

    return {
        "total_found": len(appointments),
        "processed": processed_count,
        "skipped": appointment_index - processed_count,
        "counseling": counseling_appointment_count,
    }

def main():
    # Set up the driver
    driver = setup_chrome_driver()
    
    # ping_dasrecord("Billing bot started.")
    print("\n" + "="*50)
    print("MEDICAL BILLING BOT - STARTING")
    print("="*50)
    print(f"🔧 Export Mode: {'ENABLED' if export_mode else 'DISABLED'}")
    if export_mode:
        print("📊 Mode: Export to Excel + Update Status to 'B'")
    else:
        print("📋 Mode: Normal Billing Submission")
    print("Chrome browser opened with remote debugging")
    print("If login is required, please complete it in the browser")
    
    login_success = login_to_oscar(driver)
    
    if login_success:
        print("\n" + "="*50)
        print("STARTING AUTOMATED BILLING PROCESSING")  
        print("="*50)
        
        day_sheet_window = driver.current_window_handle
        run_stats = process_appointments(driver, day_sheet_window)

        if upload_mode and export_mode:
            print("\n" + "="*50)
            print("UPLOAD MODE - Uploading export")
            print("="*50)
            upload_driver = upload_setup_driver(headless=False)
            try:
                upload_success = upload(upload_driver)
            finally:
                try:
                    upload_driver.quit()
                except Exception:
                    pass
                upload_cleanup_brave()
            stats_msg = (f"Processed: {run_stats['processed']} | Skipped: {run_stats['skipped']}"
                         + (f" | Counseling: {run_stats['counseling']}" if run_stats['counseling'] else ""))
            if upload_success:
                ping_dasrecord(f"Billing bot completed successfully. Export uploaded. {stats_msg}")
            else:
                ping_dasrecord(f"Billing bot completed. ⚠️ Upload failed — please upload manually. {stats_msg}")
        else:
            stats_msg = (f"Processed: {run_stats['processed']} | Skipped: {run_stats['skipped']}"
                         + (f" | Counseling: {run_stats['counseling']}" if run_stats['counseling'] else ""))
            ping_dasrecord(f"Billing bot completed. {stats_msg}")

        driver.quit()
        cleanup_brave()
    else:
        print("Login/setup failed - please try again")
        driver.quit()
        cleanup_brave()

if __name__ == "__main__":
    main()